import openpyxl
import urllib3
import json
import sys
"""
一覧取得Webサービス
Python3
"""

FILE_NAME = "test.xlsx"
URL = 'http://127.0.0.1:8000/api/application_inf/'
HTTP_METHOD_GET = 'GET'
HTTP_METHOD_POST = 'POST'
HTTP_METHOD_PUT = 'PUT'
COLUMNS = [
    'id', 'name', 'r_name', 'classification', 'team', 'summary', 'text', 'user'
]

PK_COLUMN = 'id'

CONTENT_TYPE = 'application/json'
AUTH_PREFIX = 'Token '
AUTH_TOKEN = '3799dec4212f96d6aa9b03827c2d08325c2d1c20'
AUTHORIZATION = AUTH_PREFIX + AUTH_TOKEN

# データ取得
def get_data():
    # 既存のファイルをロード
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active
    http = urllib3.PoolManager()
    r = http.request(HTTP_METHOD_GET, URL)

    # binaryで返ってくるため ascii にdecode
    stream = r.data.decode('ascii')
    decoder = json.JSONDecoder()

    print(stream)
    obj, index = decoder.raw_decode(stream)

    # enumerate で object と index を返す
    for row_num, row in enumerate(obj, start=1):
        for col_num, col in enumerate(COLUMNS, start=1):
            ws.cell(row=row_num, column=col_num, value=row[col])

    # 上書き保存
    wb.save(FILE_NAME)


def create_or_update(row_dict):
    # JSON エンコード
    row_json = json.dumps(row_dict, separators=(',', ':'))
    print(row_json)

    # JSON をPOST して登録
    # 更新
    if row_dict[PK_COLUMN]:
        print(URL + str(row_dict[PK_COLUMN]) + '/' )
        r = http.request(HTTP_METHOD_PUT, URL + str(row_dict[PK_COLUMN])  + '/'  ,
                headers={'Content-Type': CONTENT_TYPE,
                        'Authorization': AUTHORIZATION},
                body=row_json)
    # 登録
    else:
        print(URL + str(row_dict[PK_COLUMN]) + '/' )
        r = http.request(HTTP_METHOD_POST, URL,
                headers={'Content-Type': CONTENT_TYPE,
                        'Authorization': AUTHORIZATION},
                body=row_json)
    print(r.status, r.data)

def upload_xl_data():
    # 既存のファイルをロード
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active

    # 表全体の取得
    rows = ws.rows

    # http Request
    http = urllib3.PoolManager()

    # 初期化 1行を辞書化
    row_dict = {}

    # 表の内容を全てPOST
    # 行単位ループ
    for row in rows:
        # 列単位ループ
        for col_num, col in enumerate(row):
            row_dict[COLUMNS[col_num]] = col.value
        create_or_update(row_dict)

        # print(r.status)
        row_dict = {}

if __name__ == '__main__':
    """
    1st parameter:
        get->データの取得
        post->データの登録, 更新
    """
    args = sys.argv
    arg = str(args[1])
    print("arg:" + arg)

    if arg == "get":
        get_data()
    elif arg == "post":
        upload_xl_data()
    else:
        print("Error argument.")
        print(arg)
